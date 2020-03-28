import re
import math
import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup

# from scrapy import Selector


# url = 'http://www.plantphysiol.org'
keyword = 'proteomics'
url = f'http://www.plantphysiol.org/search/{keyword}%20jcode%3Aplantphysiol%7C%7Cplantcell%20numresults%3A'
result_num = 100
order_date = '%20sort%3Apublication-date%20direction%3Adescending'
page = 'page=1&'
year = '?facet%5Bpublication-date%5D%5B0%5D=2020'

url = f'{url}{result_num}{order_date}{year}'
# print(url)

user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'

header = {'User-Agent': user_agent}


# wb=Workbook()
# ws1 = wb.create_sheet('未筛选', 0)
# ws2 = wb.create_sheet('筛选', 1)
# ws1.column_dimensions['A'].width = 70
# ws1.column_dimensions['B'].width = 70
# ws2.column_dimensions['A'].width = 70
# ws2.column_dimensions['B'].width = 70

# title = ['名称', '链接', '摘要']
# ws1.append(title)
# ws2.append(title)

# r1 = requests.get(url, headers=header)
# result = re.search('window.googleanalytics_search_results = (.*);', r1.text).group(1)
# soup1 = BeautifulSoup(r1.text, 'html.parser')

# print(result)
# tags=soup1.find_all('a')
# for tag in tags:
# 	if tag.get('class'):
# 		linked_title = tag.get('class')[0]
# 		if linked_title == 'highwire-cite-linked-title':
# 			href = tag['href'] if tag['href'].startswith('http') else f'http://www.plantphysiol.org{tag['href']}'
# 			r2 = requests.get(href, headers=header)
# 			soup2 = BeautifulSoup(r2.text, 'html.parser')
# 			if soup2.find(attrs={"name":"DC.Description"}):
# 				name = tag.text
# 				abstract = soup2.find(attrs={"name":"DC.Description"})['content']
# 				ws1.append([name, href, abstract])
# 				if re.search('proteomics|gene|omics', abstract, re.I):
# 					ws2.append([name, href, abstract])
# 			else:
# 				ws1.append([tag.text, href])

# wb.save('C:/Users/wjb/Desktop/文献.xlsx')


class URLManager(object):
	def __init__(self):
		self.new_urls = set()
		self.old_urls = set()

	def has_new_url(self):
		# 判断是否有未爬取的url
		return self.new_url_size() != 0

	def get_new_url(self):
		# 获取一个未爬取的链接
		new_url = self.new_urls.pop()
		# 提取之后，将其添加到已爬取的链接中
		self.old_urls.add(new_url)
		return new_url

	def add_new_url(self, url):
		# 将新链接添加到未爬取的集合中(单个链接)
		if url is None:
			return
		if url not in self.new_urls and url not in self.old_urls:
			self.new_urls.add(url)

	def add_new_urls(self, urls):
		# 将新链接添加到未爬取的集合中(集合)
		if urls is None or len(urls) == 0:
			return
		for url in urls:
			self.add_new_url(url)

	def new_url_size(self):
		# 获取未爬取的url大小
		return len(self.new_urls)

	def old_url_size(self):
		# 获取已爬取的url大小
		return len(self.old_urls)


class HTMLDownload(object):
	def download(self, url):
		if url is None:
			return
		user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
		header = {'User-Agent': user_agent}
		res = requests.get(url)
		# 判断是否正常获取
		if res.status_code == 200:
			res.encoding = 'utf-8'
			res = res.text
			return res
		return None


class HTMLParser(object):

	def parser(self, page_url, html_text, keywords=list()):
		'''
		用于解析网页内容，抽取URL和数据
		:param page_url: 查询页面的URL
		:param html_text: 下载的网页内容
		:return: 返回URL和数据
		'''
		if page_url is None or html_text is None:
			return
		# soup = BeautifulSoup(html_text, 'html.parser')
		new_urls = self._get_new_urls(page_url, html_text)
		new_data1, new_data2 = self._get_new_data(page_url, html_text, keywords=keywords)
		return new_urls, new_data1, new_data2

	def _get_new_urls(self, page_url, html_text):
		'''
		抽取新的URL集合
		:param page_url:查询页面的URL
		:param soup: soup数据
		:return: 返回新的URL集合
		'''
		# page_url = 'http://www.plantphysiol.org'
		new_urls = set()
		result = int(re.search('window.googleanalytics_search_results = (.*);', html_text).group(1))
		if result > 100:
			num = math.ceil(result / 100)
			for i in range(1, num):
				page = f'page={i}&'
				url = f'{page_url}{result_num}{order_date}{page}{year}'
				new_urls.add(url)
		else:
			url = f'{page_url}{result_num}{order_date}{year}'
			new_urls.add(url)

		return new_urls

	def _get_new_data(self, page_url, html_text, keywords=list()):
		'''
		抽取有效数据
		:param page_url:查询页面的url(http://www.plantphysiol.org)
		:param soup: soup数据
		:return: 返回有效数据
		'''

		keywords = '|'.join(keywords)
		data1, data2 = [], []
		soup = BeautifulSoup(html_text, 'html.parser')
		tags = soup.find_all('a')
		for tag in tags:
			if tag.get('class'):
				linked_title = tag.get('class')[0]
				if linked_title == 'highwire-cite-linked-title':
					href = tag['href'] if tag['href'].startswith('http') else f'http://www.plantphysiol.org{tag["href"]}'
					r = requests.get(href, headers=header)
					soup2 = BeautifulSoup(r.text, 'html.parser')
					if soup2.find(attrs={"name": "DC.Description"}):
						name = tag.text
					abstract = soup2.find(attrs={"name": "DC.Description"})['content']
					data1 = [name, href, abstract]
					# ws1.append([name, href, abstract])
					if re.search('proteomics|gene|omics', abstract, re.I):
						data2 = [name, href, abstract]
					# ws2.append([name, href, abstract])
					else:
						data1 = [name, href]
					# ws1.append([tag.text, href])
		return data1, data2


class DataOutput(object):

	def __init__(self):
		self.datas1 = []
		self.datas2 = []

	def store_data(self, data1, data2):
		if data1 or data2 is None:
			return
		self.datas1.extend(data1)
		self.datas2.extend(data2)

	def output_html(self):
		wb = Workbook()
		ws1 = wb.create_sheet('未筛选', 0)
		ws2 = wb.create_sheet('筛选', 1)
		ws1.column_dimensions['A'].width = 70
		ws1.column_dimensions['B'].width = 70
		ws2.column_dimensions['A'].width = 70
		ws2.column_dimensions['B'].width = 70

		title = ['名称', '链接', '摘要']
		ws1.append(title)
		ws2.append(title)

		for data1 in self.datas1:
			ws1.append(data1)
			self.datas1.remove(data1)

		for data2 in self.datas2:
			ws2.append(data2)
			self.datas2.remove(data2)

		wb.save('C:/Users/wjb/Desktop/文献2.xlsx')


class SpiderMan(object):
	def __init__(self):
		self.manager = URLManager()
		self.downloader = HTMLDownload()
		self.parser = HTMLParser()
		self.output = DataOutput()

	def crawl(self, root_url):
		# 添加入口URL
		self.manager.add_new_url(root_url)
		# 判断url管理器中是否有新的url，同时判断抓取多少个url
		while (self.manager.has_new_url() and self.manager.old_url_size() < 100):
			try:
				# 从URL管理器获取新的URL
				new_url = self.manager.get_new_url()
				print(new_url)
				# HTML下载器下载网页
				html = self.downloader.download(new_url)
				# print(html)
				# HTML解析器抽取网页数据
				new_urls, data1, data2 = self.parser.parser(new_url, html)
				print(new_urls)
				# 将抽取的url添加到URL管理器中
				self.manager.add_new_urls(new_urls)
				# 数据存储器存储文件
				self.output.store_data(data1, data2)
				print("已经抓取%s个链接" % self.manager.old_url_size())
			except Exception as e:
				print("failed")
				print(e)
			# 数据存储器将文件输出成指定的格式
			self.output.output_html()


if __name__ == '__main__':
	# print(url)
	spider_man = SpiderMan()
	spider_man.crawl(url)

