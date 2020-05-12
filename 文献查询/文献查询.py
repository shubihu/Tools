import re
import os
# import tqdm
import math
import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup

# from scrapy import Selector

journal = int(input('请选择查询的期刊(输入对应期刊的序号即可)，\
	\n1: new phytologist(PLANT JOURNAL); \
	\n2: plant cell(plant physiology); \
	\n3: nature plant;\n:'))
terms = input('请输入要查询的内容(比如：proteomics、omics等)\n:')
keywords = input('请输入要筛选的关键词(筛选规则：摘要中若存在输入关键词中的任何一个即保留)，逗号分隔(比如：gene, omics)\n:')


user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
header = {'User-Agent': user_agent}
wb=Workbook()
ws1 = wb.create_sheet('未筛选', 0)
ws2 = wb.create_sheet('筛选', 1)
ws1.column_dimensions['A'].width = 70
ws1.column_dimensions['B'].width = 70
ws2.column_dimensions['A'].width = 70
ws2.column_dimensions['B'].width = 70

title = ['名称', '链接', '摘要']
ws1.append(title)
ws2.append(title)

def url(journal, terms):
	if journal == 1:
		url = f'https://onlinelibrary.wiley.com/action/doSearch?AfterYear=2020&AllField={terms}&BeforeYear=2020&SeriesKey=1365313x&content=articlesChapters&countTerms=true&target=default&startPage=0&sortBy=Earliest'
			    # https://onlinelibrary.wiley.com/action/doSearch?AfterYear=2020&AllField={terms}&BeforeYear=2020&SeriesKey=1365313x&content=articlesChapters&countTerms=true&sortBy=Earliest&startPage=1&target=default&pageSize=20
		return url
	elif journal == 2:
		url0 = f'http://www.plantphysiol.org/search/{terms}%20jcode%3Aplantphysiol%7C%7Cplantcell%20numresults%3A'
		result_num = 100
		order_date = '%20sort%3Apublication-date%20direction%3Adescending'
		year = '?facet%5Bpublication-date%5D%5B0%5D=2020'

		url = f'{url0}{result_num}{order_date}{year}'
		return url
	else:
		url = f'https://www.nature.com/search?q={terms}&order=date_desc&date_range=2020-2020'
		return url

# keyword = 'omics'
# url = f'https://onlinelibrary.wiley.com/action/doSearch?AfterYear=2020&AllField={keyword}&BeforeYear=2020&SeriesKey=1365313x&content=articlesChapters&countTerms=true&target=default&startPage=0&sortBy=Earliest'
def plant_journal(url, keywords):
	keywords = '|'.join(re.split('[,，]', keywords))

	r1 = requests.get(url, headers=header)
	result = re.search(r'Articles & Chapters \((\d+)\)', r1.text).group(1)
	print(f'共查询到{result}条相关文献')
	soup1 = BeautifulSoup(r1.text, 'html.parser')
	tags=soup1.find_all('a')
	for tag in tags:
		if tag.get('class'):
			publication_title = tag.get('class')[0]
			if publication_title == 'publication_title':
				href = tag['href'] if tag['href'].startswith('http') else f'https://onlinelibrary.wiley.com{tag["href"]}'
				name = tag.text
				# print(name)
				r2 = requests.get(href, headers=header)
				soup2 = BeautifulSoup(r2.text, 'html.parser')
				abstract = soup2.find('div', class_='article-section__content en main')
				if abstract:
					abstract = abstract.text
					ws1.append([name, href, abstract])

					if re.search(keywords, abstract, re.I):
						ws2.append([name, href, abstract])
				else:
					ws1.append([tag.text, href])


def plantphysiol(url, keywords):
	keywords = '|'.join(re.split('[,，]', keywords))

	url0 = f'http://www.plantphysiol.org/search/{terms}%20jcode%3Aplantphysiol%7C%7Cplantcell%20numresults%3A'
	result_num = 100
	order_date = '%20sort%3Apublication-date%20direction%3Adescending'
	year = '?facet%5Bpublication-date%5D%5B0%5D=2020'
	# url = f'{url0}{result_num}{order_date}{year}'
	r1 = requests.get(url, headers=header)
	result = int(re.search('window.googleanalytics_search_results = (.*);', r1.text).group(1))
	print(f'共查询到{result}条相关文献')
	totalPages = math.ceil(result / 100) if result > 100 else 1

	for i in range(totalPages):
		page = f'page={i}&'
		newurl = f'{url0}{result_num}{order_date}{page}{year}'
		if i == 0:
			r = r1
		else:
			r = requests.get(newurl)
		soup1 = BeautifulSoup(r.text, 'html.parser')
		tags=soup1.find_all('a')
		for tag in tags:
			if tag.get('class'):
				linked_title = tag.get('class')[0]
				if linked_title == 'highwire-cite-linked-title':
					href = tag['href'] if tag['href'].startswith('http') else f'http://www.plantphysiol.org{tag["href"]}'
					r2 = requests.get(href, headers=header)
					soup2 = BeautifulSoup(r2.text, 'html.parser')
					if soup2.find(attrs={"name":"DC.Description"}):
						name = tag.text
						abstract = soup2.find(attrs={"name":"DC.Description"})['content']
						ws1.append([name, href, abstract])
						if re.search(keywords, abstract, re.I):
							ws2.append([name, href, abstract])
					else:
						ws1.append([tag.text, href])


def nature(url, keywords):
	keywords = '|'.join(re.split('[,，]', keywords))
	r1 = requests.get(url, headers=header)
	result = re.search('"totalPages":(\d+),"page":\d+,"totalResults":(\d+)', r1.text)
	totalPages = int(result.group(1))
	totalResults = int(result.group(2))
	print(f'共查询到{totalResults}条相关文献')

	for i in range(1, totalPages + 1):
		newurl = f"{url}&page={i}"
		if i == 1:
			r = r1
		else:
			r = requests.get(newurl)
		soup1 = BeautifulSoup(r.text, 'html.parser')
		tags=soup1.find_all('a')
		for tag in tags:
			if tag.get('data-track-action') and tag.get('data-track-action') == 'search result':
				# print(tag.text)
				href = tag['href'] if tag['href'].startswith('http') else f'https://www.nature.com{tag["href"]}'
				r2 = requests.get(href, headers=header)
				soup2 = BeautifulSoup(r2.text, 'html.parser')
				if soup2.find(attrs={"name":"dc.description"}):
					abstract = soup2.find(attrs={"name":"dc.description"})['content']
					ws1.append([tag.text, href, abstract])
					if re.search(keywords, abstract, re.I):
						ws2.append([tag.text, href, abstract])
				else:
					ws1.append([tag.text, href])

if __name__ == '__main__':
	url = url(journal, terms)
	print(url)
	print('开始查询，时间可能较长, 请耐心等待\n...♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥♥...')
	if journal == 1:
		plant_journal(url, keywords)
	elif journal == 2:		
		plantphysiol(url, keywords)
	else:
		nature(url, keywords)

	wb.save(os.path.join(os.getcwd(), '文献.xlsx'))
	print(f'查询完成，结果路径为：{os.path.join(os.getcwd(), "文献.xlsx")}')
	os.system('pause')


